import pandas as pd
from conexao import conectar

def gerar_relatorio_movimentacoes():
    conn = conectar()
    try:
        query = """
        SELECT m.data_movimento AS Data, p.nome AS Produto, m.tipo_movimento AS Tipo,
               m.quantidade AS Quantidade, u.nome AS Usuario, m.observacao AS Observacao
        FROM movimentacoes m
        JOIN produtos p ON m.produto_id = p.id
        JOIN usuarios u ON m.usuario_id = u.id
        ORDER BY m.data_movimento DESC
        """
        return pd.read_sql(query, conn)
    finally:
        conn.close()

def gerar_relatorio_estoque_atual():
    conn = conectar()
    try:
        query = """
        SELECT p.nome AS Produto, p.descricao AS Descricao, p.quantidade AS Quantidade_Atual,
               p.estoque_minimo AS Estoque_Minimo, p.especificacao AS Especificacao,
               p.unidade AS Unidade, c.nome AS Categoria
        FROM produtos p
        JOIN categorias c ON p.categoria_id = c.id
        WHERE p.ativo = TRUE
        ORDER BY p.nome
        """
        return pd.read_sql(query, conn)
    finally:
        conn.close()

def gerar_relatorio_estoque_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    conn = conectar()
    try:
        query = """
        SELECT p.nome, p.descricao, p.quantidade, p.estoque_minimo,
               p.especificacao, p.unidade, c.nome AS categoria
        FROM produtos p
        JOIN categorias c ON p.categoria_id = c.id
        WHERE p.ativo = TRUE
        ORDER BY p.nome
        """
        df = pd.read_sql(query, conn)
    finally:
        conn.close()

    abrev = {'unidade': 'un', 'caixa': 'cx', 'pacote': 'pct'}
    df['unidade'] = df['unidade'].apply(lambda x: abrev.get(str(x).lower(), x) if x else x)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque Atual"

    colunas = ['Produto', 'Descrição', 'Qtd', 'Estoque Mínimo', 'Especificação', 'Un', 'Categoria']
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ws.row_dimensions[1].height = 22

    warning_fill = PatternFill("solid", fgColor="FFF3CD")

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        valores = [row.nome, row.descricao, row.quantidade, row.estoque_minimo,
                   row.especificacao, row.unidade, row.categoria]
        abaixo_minimo = row.quantidade < row.estoque_minimo

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if abaixo_minimo:
                cell.fill = warning_fill

    larguras = [30, 30, 8, 14, 30, 6, 20]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def gerar_relatorio_movimentacoes_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    df = gerar_relatorio_movimentacoes()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    colunas = list(df.columns)
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    entrada_fill = PatternFill("solid", fgColor="D4EDDA")
    saida_fill = PatternFill("solid", fgColor="F8D7DA")

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        tipo = str(row[2]).lower() if len(row) > 2 else ''
        fill = entrada_fill if tipo == 'entrada' else saida_fill
        for col in range(1, len(colunas) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[col[0].column_letter].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
